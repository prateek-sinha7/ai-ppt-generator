"""
XLSX Export Service — generates Excel workbooks from Slide_JSON using openpyxl.
"""
import io
import re
from typing import Any, Dict, List, Optional

import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# Theme color palettes (hex strings without '#' for openpyxl fills)
# ---------------------------------------------------------------------------

THEME_PALETTES: Dict[str, Dict[str, str]] = {
    "ocean-depths": {
        "primary": "1a2332",
        "secondary": "2d8b8b",
        "accent": "a8dadc",
        "text": "1a2332",
        "header_bar": "1a2332",
    },
    "sunset-boulevard": {
        "primary": "e76f51",
        "secondary": "f4a261",
        "accent": "e9c46a",
        "text": "264653",
        "header_bar": "264653",
    },
    "forest-canopy": {
        "primary": "2d4a2b",
        "secondary": "7d8471",
        "accent": "a4ac86",
        "text": "2d4a2b",
        "header_bar": "2d4a2b",
    },
    "modern-minimalist": {
        "primary": "36454f",
        "secondary": "708090",
        "accent": "d3d3d3",
        "text": "36454f",
        "header_bar": "36454f",
    },
    "golden-hour": {
        "primary": "f4a900",
        "secondary": "c1666b",
        "accent": "d4b896",
        "text": "4a403a",
        "header_bar": "4a403a",
    },
    "arctic-frost": {
        "primary": "4a6fa5",
        "secondary": "c0c0c0",
        "accent": "d4e4f7",
        "text": "2c3e50",
        "header_bar": "4a6fa5",
    },
    "desert-rose": {
        "primary": "d4a5a5",
        "secondary": "b87d6d",
        "accent": "e8d5c4",
        "text": "5d2e46",
        "header_bar": "5d2e46",
    },
    "tech-innovation": {
        "primary": "0066ff",
        "secondary": "00ffff",
        "accent": "00cccc",
        "text": "ffffff",
        "header_bar": "0066ff",
    },
    "botanical-garden": {
        "primary": "4a7c59",
        "secondary": "f9a620",
        "accent": "b7472a",
        "text": "3a3a3a",
        "header_bar": "4a7c59",
    },
    "midnight-galaxy": {
        "primary": "4a4e8f",
        "secondary": "a490c2",
        "accent": "e6e6fa",
        "text": "e6e6fa",
        "header_bar": "4a4e8f",
    },
}

_DEFAULT_THEME = "ocean-depths"

# Maximum worksheet name length per Excel spec
_MAX_SHEET_NAME = 31


def _resolve_theme(theme: str) -> Dict[str, str]:
    """Resolve a theme name to its color palette, falling back to ocean-depths."""
    key = theme.lower().replace("_", "-")
    return THEME_PALETTES.get(key, THEME_PALETTES[_DEFAULT_THEME])


def _extract_slides(slides_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extract the slides list from slides_data (dict with 'slides' key or list)."""
    if isinstance(slides_data, list):
        return slides_data
    if isinstance(slides_data, dict):
        slides = slides_data.get("slides", slides_data.get("slide_json", []))
        if isinstance(slides, list):
            return slides
    return []


def _safe_sheet_name(title: str, existing_names: set) -> str:
    """Create a valid, unique worksheet name from a slide title.

    Excel worksheet names must be ≤31 characters and cannot contain
    ``\\ / * ? : [ ]`` characters.  Duplicates get a numeric suffix.
    """
    # Strip invalid characters
    name = re.sub(r'[\\/*?\[\]:]', '', title or "Sheet").strip()
    if not name:
        name = "Sheet"
    name = name[:_MAX_SHEET_NAME]

    # Ensure uniqueness
    base = name
    counter = 2
    while name in existing_names:
        suffix = f" ({counter})"
        name = base[: _MAX_SHEET_NAME - len(suffix)] + suffix
        counter += 1

    existing_names.add(name)
    return name


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_xlsx(
    slides_data: Dict[str, Any],
    theme: str = "ocean-depths",
    design_spec: Optional[Dict] = None,
) -> bytes:
    """Generate an XLSX workbook from Slide_JSON.

    Creates worksheets for slides that contain ``chart_data`` or ``table_data``.
    If no data slides exist, a summary worksheet listing all slide titles and
    types is generated instead.

    Args:
        slides_data: Dict with a ``"slides"`` key (list of slide dicts), or a
            plain list of slide dicts.
        theme: Theme identifier (e.g. ``"ocean-depths"``).
        design_spec: Optional design specification dict (currently unused but
            reserved for future refinement).

    Returns:
        XLSX file content as ``bytes``.
    """
    palette = _resolve_theme(theme)
    slides = _extract_slides(slides_data)

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    header_fill = PatternFill(start_color=palette["header_bar"], end_color=palette["header_bar"], fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", color=palette["text"], size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="center", wrap_text=True)

    existing_names: set = set()
    data_sheets_created = 0

    for idx, slide in enumerate(slides):
        try:
            chart_data = slide.get("chart_data", slide.get("chartData", None))
            table_data = slide.get("table_data", slide.get("tableData", None))

            if chart_data and isinstance(chart_data, list):
                sheet_name = _safe_sheet_name(slide.get("title", f"Chart {idx + 1}"), existing_names)
                ws = wb.create_sheet(title=sheet_name)
                _write_chart_data(ws, chart_data, header_fill, header_font, body_font, header_align, body_align)
                data_sheets_created += 1

            if table_data and isinstance(table_data, dict):
                sheet_name = _safe_sheet_name(slide.get("title", f"Table {idx + 1}"), existing_names)
                ws = wb.create_sheet(title=sheet_name)
                _write_table_data(ws, table_data, header_fill, header_font, body_font, header_align, body_align)
                data_sheets_created += 1

        except Exception:
            logger.warning("xlsx_export.slide_render_failed", slide_index=idx, exc_info=True)

    # If no data slides, create a summary worksheet
    if data_sheets_created == 0:
        _write_summary(wb, slides, palette, header_fill, header_font, body_font, header_align, body_align)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Worksheet writers
# ---------------------------------------------------------------------------

def _write_chart_data(
    ws: Any,
    chart_data: List[Dict[str, Any]],
    header_fill: PatternFill,
    header_font: Font,
    body_font: Font,
    header_align: Alignment,
    body_align: Alignment,
) -> None:
    """Write chart data as label/value columns."""
    # Headers
    ws.cell(row=1, column=1, value="Label").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=2, value="Value").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align

    for ri, item in enumerate(chart_data, start=2):
        label = str(item.get("label", item.get("name", f"Item {ri - 1}")))
        value = item.get("value", item.get("y", 0))
        try:
            value = float(value)
        except (TypeError, ValueError):
            pass

        ws.cell(row=ri, column=1, value=label).font = body_font
        ws.cell(row=ri, column=1).alignment = body_align
        ws.cell(row=ri, column=2, value=value).font = body_font
        ws.cell(row=ri, column=2).alignment = body_align

    # Auto-size columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def _write_table_data(
    ws: Any,
    table_data: Dict[str, Any],
    header_fill: PatternFill,
    header_font: Font,
    body_font: Font,
    header_align: Alignment,
    body_align: Alignment,
) -> None:
    """Write table data with headers and rows."""
    headers = table_data.get("headers", [])
    rows = table_data.get("rows", [])

    if not headers and not rows:
        return

    # Write headers
    for ci, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=str(header_text))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data rows
    start_row = 2 if headers else 1
    for ri, row in enumerate(rows, start=start_row):
        if isinstance(row, dict):
            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(row=ri, column=ci, value=str(row.get(h, "")))
                cell.font = body_font
                cell.alignment = body_align
        elif isinstance(row, (list, tuple)):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=str(val))
                cell.font = body_font
                cell.alignment = body_align

    # Auto-size columns (approximate)
    for ci in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[col_letter].width = max(14, min(40, len(str(headers[ci - 1])) + 6)) if ci <= len(headers) else 14


def _write_summary(
    wb: Workbook,
    slides: List[Dict[str, Any]],
    palette: Dict[str, str],
    header_fill: PatternFill,
    header_font: Font,
    body_font: Font,
    header_align: Alignment,
    body_align: Alignment,
) -> None:
    """Write a summary worksheet listing all slide titles and types."""
    ws = wb.create_sheet(title="Summary")

    # Headers
    for ci, header_text in enumerate(["#", "Title", "Type"], start=1):
        cell = ws.cell(row=1, column=ci, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for ri, slide in enumerate(slides, start=2):
        ws.cell(row=ri, column=1, value=ri - 1).font = body_font
        ws.cell(row=ri, column=1).alignment = body_align
        ws.cell(row=ri, column=2, value=slide.get("title", "Untitled")).font = body_font
        ws.cell(row=ri, column=2).alignment = body_align
        ws.cell(row=ri, column=3, value=slide.get("type", "content")).font = body_font
        ws.cell(row=ri, column=3).alignment = body_align

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18

    if not slides:
        ws.cell(row=2, column=1, value="No slides in presentation.").font = body_font
